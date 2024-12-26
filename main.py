from flask import Flask, request, jsonify, render_template, send_file, Response, send_from_directory
from flask_cors import CORS
from db import VisitorDB
import csv
from io import StringIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from openpyxl.utils import get_column_letter
import json
import time
from cachetools import TTLCache, LRUCache
from collections import defaultdict
import threading
from queue import Queue, Empty
import os
from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__, 
           static_url_path='/visit/static',  # static URL 경로
           static_folder='static')           # static 폴더 위치

app.config['APPLICATION_ROOT'] = '/visit'
app.wsgi_app = ProxyFix(app.wsgi_app)

CORS(app)
db = VisitorDB()

# 디버그 로그 추가
print(f"Database initialized at: {db.db_path}")
print(f"Static URL path: {app.static_url_path}")
print(f"Static folder: {app.static_folder}")

# 더 큰 캐시와 긴 TTL
visitors_cache = TTLCache(maxsize=100, ttl=5)  # 5초 TTL
analytics_cache = LRUCache(maxsize=50)  # LRU 캐시 추가

# 클라이언트 연결 관리
clients = defaultdict(list)
client_lock = threading.Lock()

def notify_clients(data):
    with client_lock:
        for client_list in clients.values():
            for client in client_list:
                client.put(data)

def adjust_column_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            
            # 한글 문자 길이 보정
            korean_char_count = sum(1 for c in str(cell.value) if ord('가') <= ord(c) <= ord('힣'))
            max_length += korean_char_count * 0.5
        
        # 최소 너비 설정
        adjusted_width = max(max_length + 2, 10)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# URL prefix 처리를 위한 함수
def get_prefix():
    return app.config['APPLICATION_ROOT']

@app.route('/')
def index():
    print(f"Serving index page from {get_prefix()}")  # 디버그 로그
    return render_template('index.html')

@app.route('/api/visitors', methods=['POST'])
def add_visitor():
    data = request.json
    if not data:
        return jsonify({
            'status': 'error',
            'message': '요청 데이터가 없습니다.'
        }), 400

    required_fields = ['company', 'name', 'position', 'visit_location', 'visit_purpose', 'manager']
    for field in required_fields:
        if field not in data:
            return jsonify({
                'status': 'error',
                'message': f'{field} 필드가 필요합니다.'
            }), 400

    print(f"Received visitor data: {data}")  # 디버그 로그
    
    visitor_id = db.add_visitor(
        company=data['company'],
        name=data['name'],
        position=data['position'],
        contact=data.get('contact', ''),  # 선택적 필드
        visit_location=data['visit_location'],
        visit_purpose=data['visit_purpose'],
        manager=data['manager']
    )
    
    if visitor_id == -1:
        print("Duplicate visitor rejected")  # 거부 로그
        return jsonify({
            'status': 'error',
            'message': '이미 입실한 방문자입니다.'
        }), 400
    
    print(f"Visitor added successfully with id: {visitor_id}")  # 성공 로그
    visitors = db.get_current_visitors()
    update_visitors(visitors)  # 변경 시에만 알림
    return jsonify({'id': visitor_id}), 201

@app.route('/api/visitors/<int:visitor_id>/checkout', methods=['POST'])
def checkout_visitor(visitor_id):
    db.check_out_visitor(visitor_id)
    return jsonify({'status': 'success'}), 200

@app.route('/api/current-visitors', methods=['GET'])
def get_current_visitors():
    try:
        visitors = db.get_current_visitors()
        print(f"Retrieved visitors: {visitors}")  # 디버그 로그
        return jsonify(visitors)
    except Exception as e:
        print(f"Error getting visitors: {e}")  # 에러 로그
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats/managers', methods=['GET'])
def get_manager_stats():
    stats = db.get_manager_stats()
    return jsonify(stats)

@app.route('/api/stats/companies', methods=['GET'])
def get_company_stats():
    stats = db.get_company_stats()
    return jsonify(stats)

@app.route('/api/visitors/<date>', methods=['GET'])
def get_visitors_by_date(date):
    try:
        # 날짜 형식 검증
        datetime.strptime(date, '%Y-%m-%d')
        visitors = db.get_visitors_by_date(date)
        print(f"Retrieved visitors for {date}: {visitors}")  # 디버그 로그
        return jsonify(visitors)
    except ValueError:
        return jsonify({'error': '잘못된 날짜 형식입니다.'}), 400
    except Exception as e:
        print(f"Error getting visitors for {date}: {e}")  # 에러 로그
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/<int:year>/<int:month>', methods=['GET'])
def export_excel(year, month):
    visitors = db.get_visitors_by_month(year, month)
    missed_checkouts = db.get_missed_checkouts_by_month(year, month)
    
    wb = Workbook()
    
    # 방문 기록 시트
    ws1 = wb.active
    ws1.title = "방문 기록"
    
    # 헤더 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # 방문 기록 헤더
    headers = ['날짜', '업체명', '성명', '직급', '연락처', '방문장소', 
              '방문목적', '입실시간', '퇴실시간', '담당자', '상태']
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 방문 기록 데이터
    for row, v in enumerate(visitors, 2):
        status = '정상' if v[9] else '미퇴실'
        data = [v[1], v[2], v[3], v[4], v[5], v[6], v[7], v[8], 
                v[9] or '-', v[10], status]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
    
    # 열 너비 자동 조정
    adjust_column_width(ws1)
    
    # 퇴실 누락 기록 시트
    ws2 = wb.create_sheet(title="퇴실 누락 기록")
    
    # 퇴실 누락 헤더
    missed_headers = ['날짜', '업체명', '성명', '직급', '방문장소', '입실시간', 
                     '처리일자', '처리사유']
    
    for col, header in enumerate(missed_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # 퇴실 누락 데이터
    for row, m in enumerate(missed_checkouts, 2):
        data = [m[6], m[1], m[2], m[3], m[4], m[5], m[7], m[8]]
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
    
    # 열 너비 자동 조정
    adjust_column_width(ws2)
    
    # 파일 저장
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'방문기록_{year}_{month}.xlsx'
    )

@app.route('/api/analytics/companies', methods=['GET'])
def get_company_analytics():
    analytics = db.get_company_analytics()
    return jsonify(analytics)

@app.route('/api/analytics/purposes', methods=['GET'])
def get_purpose_ranking():
    ranking = db.get_purpose_ranking()
    return jsonify(ranking)

@app.route('/api/options/companies', methods=['GET'])
def get_companies():
    try:
        companies = db.get_companies()
        print(f"Companies API response: {companies}")  # 디버그 로그
        return jsonify(companies)
    except Exception as e:
        print(f"Error in companies API: {str(e)}")  # 에러 로그
        return jsonify({'error': str(e)}), 500

@app.route('/api/options/positions', methods=['GET'])
def get_positions():
    return jsonify(db.get_positions())

@app.route('/api/options/locations', methods=['GET'])
def get_locations():
    return jsonify(db.get_locations())

@app.route('/api/options/departments', methods=['GET'])
def get_departments():
    departments = db.get_departments()
    print("Departments:", departments)
    return jsonify(departments)

@app.route('/api/managers/search', methods=['GET'])
def search_managers():
    query = request.args.get('q', '')
    return jsonify(db.search_managers(query))

@app.route('/api/managers/department/<int:dept_id>', methods=['GET'])
def get_managers_by_department(dept_id):
    managers = db.get_managers_by_department(dept_id)
    print(f"Managers for department {dept_id}:", managers)
    return jsonify(managers)

@app.route('/api/visitor-history', methods=['GET'])
def get_visitor_history():
    company = request.args.get('company')
    name = request.args.get('name')
    return jsonify(db.get_visitor_history(company, name))

@app.route('/api/options/purposes', methods=['GET'])
def get_visit_purposes():
    return jsonify(db.get_visit_purposes())

@app.route('/api/options/companies', methods=['POST'])
def add_company():
    data = request.json
    success, message = db.add_company(data['name'])
    if success:
        return jsonify({'status': 'success', 'message': message}), 201
    else:
        return jsonify({'status': 'error', 'message': message}), 400

@app.route('/api/missed-checkouts', methods=['GET'])
def get_missed_checkouts():
    return jsonify(db.get_missed_checkouts())

@app.route('/api/visitors/<int:visitor_id>/missed-checkout', methods=['POST'])
def mark_missed_checkout(visitor_id):
    data = request.json
    db.add_missed_checkout(visitor_id, data['original_date'], data['reason'])
    return jsonify({'status': 'success'}), 200

@app.route('/api/visitors/check-duplicate', methods=['POST'])
def check_duplicate_visitor():
    data = request.json
    print(f"Checking duplicate for: {data}")  # 디버그 로그
    
    # 기존 방문자 조회
    conn = db.get_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, company, name, position, check_in_time 
            FROM visitors 
            WHERE date = ? 
            AND company = ? 
            AND name = ? 
            AND position = ? 
            AND check_out_time IS NULL
        ''', (
            datetime.now().strftime('%Y-%m-%d'),
            data['company'],
            data['name'],
            data['position']
        ))
        existing_visitor = cursor.fetchone()
        
        if existing_visitor:
            return jsonify({
                'isDuplicate': True,
                'existingVisitor': {
                    'id': existing_visitor[0],
                    'company': existing_visitor[1],
                    'name': existing_visitor[2],
                    'position': existing_visitor[3],
                    'check_in_time': existing_visitor[4]
                }
            })
        
        return jsonify({'isDuplicate': False})
        
    finally:
        conn.close()

@app.route('/mobile-register')
def mobile_register():
    print(f"Serving mobile register page from {get_prefix()}")  # 디버그 로그
    return render_template('mobile-register.html')

@app.route('/qr')
def qr_code():
    mobile_url = request.host_url.rstrip('/') + url_for('mobile_register')
    print(f"QR code URL: {mobile_url}")  # 디버그 로그
    return render_template('qr.html', mobile_url=mobile_url)

@app.route('/api/sse')
def sse():
    client_id = request.remote_addr
    queue = Queue()
    
    with client_lock:
        clients[client_id].append(queue)
    
    def generate():
        try:
            while True:
                data = queue.get(timeout=30)  # 30초 타임아웃
                yield f"data: {data}\n\n"
        except Empty:
            pass
        finally:
            with client_lock:
                clients[client_id].remove(queue)
                if not clients[client_id]:
                    del clients[client_id]
    
    return Response(generate(), mimetype='text/event-stream')

# 데이터 변경 시에만 알림
def update_visitors(visitor_data):
    data = json.dumps(visitor_data)
    notify_clients(data)

def get_cached_visitors():
    cache_key = 'current_visitors'
    if cache_key not in visitors_cache:
        visitors = db.get_current_visitors()
        visitors_cache[cache_key] = visitors
    return visitors_cache[cache_key]

# 이미지 파일 서빙을 위한 라우트 추가
@app.route('/static/images/<path:filename>')
def serve_image(filename):
    return send_from_directory('static/images', filename)

@app.route('/api/db-test')
def test_db():
    try:
        # 데이터베이스 연결 테스트
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 테이블 존재 확인
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        
        # 방문자 수 확인
        cursor.execute("SELECT COUNT(*) FROM visitors")
        visitor_count = cursor.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'tables': tables,
            'visitor_count': visitor_count,
            'db_path': db.db_path
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'db_path': db.db_path
        }), 500

# 정적 파일 처리
@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

@app.route('/visit/debug/db-test')
def test_db_connection():
    try:
        # 데이터베이스 파일 존재 확인
        print(f"DB Path: {db.db_path}")
        print(f"DB File exists: {os.path.exists(db.db_path)}")
        
        # 테이블 목록 조회
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        
        # 각 테이블의 레코드 수 확인
        table_counts = {}
        for table in tables:
            cursor.execute(f"SELECT COUNT(*) FROM {table[0]}")
            count = cursor.fetchone()[0]
            table_counts[table[0]] = count
        
        conn.close()
        
        return jsonify({
            'status': 'success',
            'db_path': db.db_path,
            'tables': [t[0] for t in tables],
            'record_counts': table_counts
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'db_path': db.db_path
        }), 500

if __name__ == '__main__':
    app.run(debug=True)
